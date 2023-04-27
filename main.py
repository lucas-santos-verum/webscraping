

from datetime import datetime
import datetime
import streamlit as st
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import locale
import openai
import pandas as pd
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, Color
import os


openai.api_key = st.secrets['api_key']

locale.setlocale(locale.LC_ALL, 'pt_BR')


def run(words, n_links, date_cut, tribo=None, automated=False):
    options = Options()
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--start-maximized')
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    # options.headless = True
    webdriver_service = Service(r'C:\Users\LucasSantos\Documents\webscraping_final\chromedriver\chromedriver.exe')
    driver = webdriver.Chrome(executable_path=r'C:\Users\LucasSantos\Documents\webscraping_final\chromedriver\chromedriver.exe', options=options)
    driver.get('https://www.google.com.br/')
    try:
        text_area = driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/form/div[1]/div[1]/div[1]/div/div[2]/input')
    except:
        try:
            text_area = driver.find_element(By.XPATH, '//*[@id="APjFqb"]')
        except:
            pass
    text_area.send_keys(words)
    text_area.send_keys(Keys.ENTER)
    links = []
    com_data = []
    dados = []
    driver.set_page_load_timeout(30)


    while len(links) < n_links:
        # pega todos os links da página
        folders = driver.find_elements(By.CLASS_NAME, 'yuRUbf')

        # se tiver data, adiciona a lista com_data
        for i in range(len(folders)):
            try:
                com_data.append(
                    driver.find_element(By.XPATH, '//*[@id="rso"]/div[{}]/div/div/div[2]/div/span[1]'.format(i)))
            except:
                pass

        # adiciona o link e a data de cada noticia na lista com_data à lista links
        for j in range(len(com_data)):
            if len(links) < n_links:
                try:
                    links.append((com_data[j].find_element(By.XPATH,
                                                           './/ancestor::div[2]//preceding-sibling::div/div/a').get_attribute(
                        'href'), com_data[j].find_element(By.XPATH, './span').text))
                except:
                    pass

        if len(links) < n_links:
            print('TESTE')
            try:
                driver.get(
                    driver.find_element(By.XPATH, '//*[@id="pnnext"]').get_attribute(
                        'href')
                )
            except:
                driver.get(
                    driver.find_element(By.XPATH, '/html/body/div[7]/div/div[11]/div/div[4]/div/div[2]/table/tbody/tr/td[12]/a').get_attribute(
                        'href')
                )
        else:
            driver.close()

    for l in links:
        try:
            if datetime.datetime.strptime(l[1], '%d de %b. de %Y').date() > date_cut:
                result_list = ask_chatGPT(l[0], automated, tribo)
                dados.append(result_list)
        except:
            pass
    if not automated:
        df = pd.DataFrame(dados, columns=['Link', 'Sumário', 'Empresa', 'Investimento', 'Local', 'Contato', 'Cold Call'])
        report_toExcel(df, 'Report')
        with open('Report.xlsx', 'rb') as my_file:
            st.download_button(label='Download', data=my_file, file_name='report.xlsx')
    else:
        df = pd.DataFrame(dados,
                          columns=['Tribo','Link', 'Sumário', 'Empresa', 'Investimento', 'Local', 'Contato', 'Cold Call'])
        report_toExcel(df, 'ReportAutomatico')
        with open('ReportAutomatico.xlsx', 'rb') as my_file:
            st.download_button(label='Download', data=my_file, file_name='reportautomatico.xlsx')



def ask_chatGPT(link, automated, tribo=None):

    response = openai.Completion.create(model="text-davinci-003",
                                        prompt="Sumarize a notícia deste link:\n\n " + link, temperature=0.50,
                                        max_tokens=800, top_p=1.0, frequency_penalty=0.8, presence_penalty=0.0)
    q1 = response['choices'][0]['text']

    response = openai.Completion.create(model="text-davinci-003",
                                        prompt="Qual empresa (escreva somente o nome da empresa, e não pode ser o "
                                               "site da notícia nem uma área muito abragente, por exemplo, "
                                               "saneamento básico, se não tiver uma empresa específica, diga que não "
                                               "tem) é o foco"
                                               "desta notícia:\n\n " + link,
                                        temperature=0.25,
                                        max_tokens=100, top_p=1.0, frequency_penalty=0.8, presence_penalty=0.0)
    q2 = response['choices'][0]['text']

    response = openai.Completion.create(model="text-davinci-003",
                                        prompt="Qual o valor a ser investido (responda apenas o valor, em milhões ou bilhões) pela empresa foco desta notícia:\n\n " + link,
                                        temperature=0.25,
                                        max_tokens=100, top_p=1.0, frequency_penalty=0.8, presence_penalty=0.0)
    q3 = response['choices'][0]['text']

    response = openai.Completion.create(model="text-davinci-003",
                                        prompt="Quais as principais localidades são mencionadas neste link:\n\n " + link,
                                        temperature=0.25,
                                        max_tokens=150, top_p=1.0, frequency_penalty=0.8, presence_penalty=0.0)
    q4 = response['choices'][0]['text']

    response = openai.Completion.create(model="text-davinci-003",
                                        prompt="Indique o gestor/gerente/CEO/CFO/Diretor/Executivo mencionado na notícia:\n\n " + link,
                                        temperature=0.75,
                                        max_tokens=150, top_p=1.0, frequency_penalty=0.8, presence_penalty=0.0)
    q5 = response['choices'][0]['text']

    response = openai.Completion.create(model="text-davinci-003",
                                        prompt="Imagine que você é da empresa Verum Partners (uma empresa de "
                                               "engenharia que oferece serviços como Gestão Integrada de Projetos, "
                                               "Transformação Digital, Planejamento de Implantação, etc) e quer "
                                               "conseguir uma"
                                               "parceria com a principal empresa mencionada no link. Monte um texto "
                                               "de cold call com base na notícia:\n\n "
                                               + link + "Quero que você utilize o seguinte modelo: Sou da Verum "
                                                        "Partners, uma empresa de consultoria atuante no mercado de "
                                                        "infraestrutura, construção e gerenciamento de projetos. "
                                                        "Soubemos que a [nome da empresa] irá [informar o que a "
                                                        "empresa irá fazer, investir, construir, ampliar, modernizar, "
                                                        "etc.] o seu[informar o que será o investimento, uma planta, "
                                                        "um polo, uma fábrica, etc.]. Este [informar a fonte da "
                                                        "informação, notícia, anuncio, etc.] nos chamou atenção "
                                                        "porque o investimento contempla a realização de [informar "
                                                        "porque tem relação com nosso negócio], desafio muito similar "
                                                        "ao que temos nos envolvido e obtido resultados relevantes "
                                                        "junto à nossos clientes. Pensando em ajudá-los a posicionar "
                                                        "este [informar o que será projeto, operação, transação, "
                                                        "etc.] como um case de sucesso, gostaria de agendar uma "
                                                        "rápida reunião contigo para apresentar como ajudamos outras "
                                                        "organizações como a [empresa do cliente] em projetos "
                                                        "similares, compreender os seus desafios, bem como analisar "
                                                        "se podemos trabalhar juntos."
                                                        ""
                                                        "Substitua tudo que está entre parenteses pelas informações adequadas.",
                                        temperature=0.50,
                                        max_tokens=400, top_p=1.0, frequency_penalty=0.8, presence_penalty=0.0)
    q6 = response['choices'][0]['text']

    if not automated:
        return [link, q1, q2, q3, q4, q5, q6]
    return [tribo, link, q1, q2, q3, q4, q5, q6]

def report_toExcel(df, file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório'
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 50
    thin = Side(border_style="thin", color="000000")
    cols_name = ['Link', 'Sumário', 'Empresa', 'Investimento', 'Local', 'Contato', 'Cold Call']
    col_count = 1
    for name in cols_name:
        cell = ws.cell(row=1, column=col_count)
        col_count += 1
        cell.value = name
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size="12", color='FFFFFF')
        cell.fill = PatternFill(start_color="044F83", end_color="044F83", fill_type="solid")

    count_row, count_col = 2, 1
    for row in df.index:
        for col in df.columns:
            cell = ws.cell(row=count_row, column=count_col)
            cell.value = df.at[row, col]
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            cell.font = Font(bold=False, size = "12", color='404040')
            cell.fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type="solid")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            count_col += 1
        count_row += 1
        count_col = 1
    wb.save(file_name + ".xlsx")

def report_toExcel_automated(df, file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório'
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 50
    thin = Side(border_style="thin", color="000000")
    cols_name = ['Tribo','Link', 'Sumário', 'Empresa', 'Investimento', 'Local', 'Contato', 'Cold Call']
    col_count = 1
    for name in cols_name:
        cell = ws.cell(row=1, column=col_count)
        col_count += 1
        cell.value = name
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size="12", color='FFFFFF')
        cell.fill = PatternFill(start_color="044F83", end_color="044F83", fill_type="solid")

    count_row, count_col = 2, 1
    for row in df.index:
        for col in df.columns:
            cell = ws.cell(row=count_row, column=count_col)
            cell.value = df.at[row, col]
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            cell.font = Font(bold=False, size = "12", color='404040')
            cell.fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type="solid")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            count_col += 1
        count_row += 1
        count_col = 1
    wb.save(file_name + ".xlsx")

def automated():
    workbook = openpyxl.load_workbook(filename='Modelo de Input - BD.xlsx')
    sheet = workbook['Input']
    # nova_planilha = workbook.create_sheet(title='Resultado')
    print('Teste')

    for row in sheet.iter_rows(min_row=2, values_only=True):  # começa na segunda linha
        # Extrai as palavras-chave da linha (excluindo a primeira coluna)
        words = [word for word in row[1:3] if word is not None]
        # Extrai a data de corte da linha
        data_corte = row[4]  # assumindo que a data de corte está na quinta coluna
        # Extrai o número máximo de notícias desejados da linha
        numero_noticias = row[3]  # assumindo que o número máximo de notícias está na quarta coluna
        # Extrai a tribo da linha
        tribo = row[0]  # assumindo que a tribo está na primeira coluna
        # Chama a função run com as palavras-chave, data de corte e número máximo de notícias da linha
        run(words, numero_noticias, data_corte, tribo, True)
        # Adiciona as informações extraídas em um dataframe para posterior processamento
        # df = df.append({'Tribo': tribo, 'Palavras-chave': ', '.join(words), 'Data de corte': data_corte,
        #                'Número máximo de notícias': numero_noticias}, ignore_index=True)

st.title('Verum Partners - Clipping')

with st.expander('O que é?'):
    st.write('Ferramenta de ***Clipping*** desenvolvida pela Verum Partners para auxiliar na identificação de novas '
             'oportunidades comerciais.')
with st.expander('Como funciona?'):
    st.write('A ferramenta busca notícias atuais e relevantes sobre investimentos nos setores e empresas de sua '
             'escolha, gerando um relatório com informações sobre o investimento e um texto para ajudar no contato '
             'com a empresa.')
with st.expander('Como usar?'):
    st.write('Preencha os campos de setor e empresa (caso não tenha uma empresa específica em mente, pode deixar em '
             'branco), selecione uma data de corte (serão selecionadas somente notícias a partir dessa data), '
             'selecione um número máximo de notícias e clique em Buscar. No final, baixe o relatório gerado.')

with st.sidebar:
    # keywords = st.text_input('Palavras-Chave (;)', 'Mineração; Investimentos')
    setor = st.text_input('Setor', 'Saneamento')
    empresa = st.text_input('Empresa', 'Iguá')
    date_cut = st.date_input(
        'Data de Corte',
        datetime.date.today() - datetime.timedelta(days=90))
    st.caption('Serão coletadas notícias apenas a partir da data escolhida')
    n_links = st.slider('Nº Máximo de Notícias', 5, 30, 10)
    st.caption('As notícias são ordenadas de acordo com os resultados do Google, ou seja, por relevância.')
    btn_search = st.button('Buscar')

if btn_search:
    # words = ' '.join(keywords.strip().replace(' ', '').split(';'))
    words = 'Investimentos'+' ' + setor + ' ' + empresa
    numCols = 0
    with st.spinner('Aguarde, isso pode levar alguns minutos'):
        run(words, n_links, date_cut)





